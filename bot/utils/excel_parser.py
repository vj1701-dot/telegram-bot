"""Parse schedule files - supports multiple .xlsx files."""
import logging
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from typing import List

from shared.models import ScheduleEntry

logger = logging.getLogger(__name__)


class ExcelParser:
    """Parses Excel schedule files - supports multiple schedules."""

    def __init__(self, data_dir: Path = Path("/app/data")):
        self.data_dir = data_dir

    def get_all_schedule_files(self) -> List[Path]:
        """Get all schedule.xlsx files in data directory."""
        schedule_files = list(self.data_dir.glob("schedule*.xlsx"))
        logger.info(f"Found {len(schedule_files)} schedule file(s): {[f.name for f in schedule_files]}")
        return schedule_files

    def get_all_entries(self) -> List[ScheduleEntry]:
        """Get all schedule entries from ALL Excel files."""
        all_entries = []
        schedule_files = self.get_all_schedule_files()

        if not schedule_files:
            logger.warning("No schedule files found")
            return []

        for excel_path in schedule_files:
            entries = self._parse_file(excel_path)
            all_entries.extend(entries)
            logger.info(f"Loaded {len(entries)} entries from {excel_path.name}")

        logger.info(f"Total: {len(all_entries)} entries from {len(schedule_files)} file(s)")
        return all_entries

    def _parse_file(self, excel_path: Path) -> List[ScheduleEntry]:
        """Parse a single Excel file."""
        if not excel_path.exists():
            logger.warning(f"Schedule file not found: {excel_path}")
            return []

        try:
            wb = load_workbook(excel_path, read_only=True)
            ws = wb.active

            entries = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                entry = self._parse_row(row, row_idx)
                if entry:
                    entries.append(entry)

            wb.close()
            return entries

        except Exception as e:
            logger.error(f"Failed to parse {excel_path.name}: {e}")
            return []

    def get_today_entries(self) -> List[ScheduleEntry]:
        """Get entries scheduled for today."""
        today = datetime.now().strftime("%Y-%m-%d")
        all_entries = self.get_all_entries()

        today_entries = [e for e in all_entries
                        if e.date == today and e.enabled]

        logger.info(f"Found {len(today_entries)} entries for today ({today})")
        return today_entries

    def get_entries_by_date(self, date_str: str) -> List[ScheduleEntry]:
        """Get entries for a specific date (YYYY-MM-DD format)."""
        all_entries = self.get_all_entries()
        return [e for e in all_entries if e.date == date_str and e.enabled]

    def _parse_row(self, row, row_idx: int) -> ScheduleEntry:
        """Parse a single Excel row into ScheduleEntry."""
        # Expected columns: Date, Path, Track Name, Enabled
        try:
            if not row or len(row) < 3:
                return None

            date_cell = row[0]
            path_cell = row[1]
            track_cell = row[2]
            enabled_cell = row[3] if len(row) > 3 else True

            if not all([date_cell, path_cell, track_cell]):
                return None

            # Handle datetime objects from Excel
            if isinstance(date_cell, datetime):
                date_str = date_cell.strftime("%Y-%m-%d")
            else:
                date_str = str(date_cell)

            return ScheduleEntry(
                date=date_str,
                path=str(path_cell),
                track_name=str(track_cell),
                enabled=bool(enabled_cell)
            )
        except Exception as e:
            logger.error(f"Failed to parse row {row_idx}: {e}")
            return None
